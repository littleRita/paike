# -*- coding: utf-8 -*-
"""把求解结果写入 今年最终课表_待填写.xlsx，格式照抠去年真实 最终课表.xls 的版式。"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

from parse_input import load_classes, apply_confirmed_disambiguations, BASE

OUTPUT_XLSX = BASE / "今年最终课表_待填写.xlsx"

DAY_NAMES = ["星期一", "星期二", "星期三", "星期四", "星期五"]
DAY_COLS = {1: 4, 2: 5, 3: 6, 4: 7, 5: 8}  # D,E,F,G,H
GRADE26 = ("2026级", "2025级")


def periods_for_grade_row_layout(grade_key):
    """返回该年级最终课表需要几行"节"（一二年级6节；三到六年级7节，但第7节只在周二有内容）。"""
    return 6 if grade_key in GRADE26 else 7


def format_atom_lines(atom):
    """把一个atom格式化成(科目行文字, 教师行文字)。
    combo科目按去年真实样例的确切格式（见一年级1班周四第5节）：
    第一行(通常是"科目行"这一行) = "科目A(教师A)"；第二行(通常是"教师行"这一行) = "科目B(教师B)"。
    """
    if atom["kind"] == "combo":
        subj_a, subj_b = atom["subject"].split("/")
        teacher_a, teacher_b = atom["teachers"]
        return f"{subj_a}({teacher_a})", f"{subj_b}({teacher_b})"
    subj = atom["subject"]
    teacher = ",".join(atom["teachers"])
    return subj, f"({teacher})"


def build_class_sheet(wb, c, atoms, solution_for_class, used_names):
    grade_key = c["grade_key"]
    base_name = f"{c['grade_name']}{c['class_no']}班 班主任_{c['homeroom_raw'] or ''}"
    sheet_name = base_name[:31]
    n = 1
    while sheet_name in used_names:
        n += 1
        sheet_name = f"{base_name[:28]}_{n}"
    used_names.add(sheet_name)
    ws = wb.create_sheet(title=sheet_name)

    max_periods = periods_for_grade_row_layout(grade_key)

    grid = {}
    for idx, slot in solution_for_class.items():
        day, period = slot // 10, slot % 10
        grid[(day, period)] = atoms[idx]

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.cell(row=1, column=1, value="2026-2027学年班级课程表")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=2, column=8, value=f"{c['grade_name']}{c['class_no']}班 班主任:{c['homeroom_raw'] or ''}")

    ws.cell(row=3, column=1, value="　　　　星期\n\n　时间")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    for day, col in DAY_COLS.items():
        ws.cell(row=3, column=col, value=DAY_NAMES[day - 1]).alignment = center

    note_row = 4
    ws.cell(row=note_row, column=2, value="自主阅读+晨会+体育活动")
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=8)

    period_names = ["第一节", "第二节", "第三节", "第四节", "第五节", "第六节", "第七节"][:max_periods]
    row = note_row + 1
    am_start_row = note_row
    am_end_row = None
    pm_start_row = None
    for p_idx, pname in enumerate(period_names, start=1):
        ws.cell(row=row, column=2, value=pname)
        ws.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=2)
        for day, col in DAY_COLS.items():
            atom = grid.get((day, p_idx))
            if atom:
                subj_line, teacher_line = format_atom_lines(atom)
                ws.cell(row=row, column=col, value=subj_line).alignment = center
                ws.cell(row=row + 1, column=col, value=teacher_line).alignment = center
        if p_idx == 4:
            am_end_row = row + 1
            pm_start_row = row + 2
        row += 2

    ws.merge_cells(start_row=am_start_row, start_column=1, end_row=am_end_row, end_column=1)
    ws.cell(row=am_start_row, column=1, value="上\n\n午").alignment = center
    ws.merge_cells(start_row=pm_start_row, start_column=1, end_row=row - 1, end_column=1)
    ws.cell(row=pm_start_row, column=1, value="下\n\n午").alignment = center

    ws.cell(row=row, column=2, value="课后服务")

    for col_letter, width in (("A", 6), ("B", 10), ("C", 3)):
        ws.column_dimensions[col_letter].width = width
    for col in "DEFGH":
        ws.column_dimensions[col].width = 16

    return ws


def write_workbook(classes, all_atoms, solution):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_names = set()
    missing = []
    for c in classes:
        key = (c["grade_key"], c["class_no"])
        if key not in solution:
            missing.append(key)
            continue
        build_class_sheet(wb, c, all_atoms[key], solution[key], used_names)
    wb.save(OUTPUT_XLSX)
    return missing


if __name__ == "__main__":
    classes = load_classes()
    apply_confirmed_disambiguations(classes)
    from build_atoms import build_all_atoms
    all_atoms, _ = build_all_atoms(classes)

    sol_path = BASE / "_solution.json"
    with open(sol_path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    solution = {}
    for k, v in raw.items():
        g, n = k.split("|")
        solution[(g, int(n))] = {int(i): slot for i, slot in v.items()}

    missing = write_workbook(classes, all_atoms, solution)
    if missing:
        print(f"警告：{len(missing)}个班没有解，未写入: {missing}")
    print(f"已写出 {OUTPUT_XLSX}")
