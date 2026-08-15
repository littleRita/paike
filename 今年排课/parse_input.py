# -*- coding: utf-8 -*-
"""解析《教师与课程安排.xlsx》为结构化数据，供排课求解器使用。"""
import re
import json
from pathlib import Path
from collections import defaultdict

import openpyxl

BASE = Path(__file__).parent
INPUT_XLSX = BASE / "教师与课程安排.xlsx"

GRADE_NAME = {
    "2026级": "一年级", "2025级": "二年级", "2024级": "三年级",
    "2023级": "四年级", "2022级": "五年级", "2021级": "六年级",
}
GRADE_ORDER = ["2026级", "2025级", "2024级", "2023级", "2022级", "2021级"]
CAMPUS = {"2026级": "B", "2025级": "A", "2024级": "A", "2023级": "A", "2022级": "A", "2021级": "A"}

# 每周节次结构：day -> 总节数；1-2年级周二~周五只到第5节，周一到第6节
PERIOD_STRUCTURE_26 = {1: 6, 2: 5, 3: 5, 4: 5, 5: 5}   # 一二年级：26节/周
PERIOD_STRUCTURE_31 = {1: 6, 2: 7, 3: 6, 4: 6, 5: 6}   # 三到六年级：31节/周
AM_PERIODS = {1, 2, 3, 4}


def periods_for_grade(grade_key):
    return PERIOD_STRUCTURE_26 if grade_key in ("2026级", "2025级") else PERIOD_STRUCTURE_31


def parse_header_segment_text(header_text):
    """把列头文本（如"语文6+道德与法治1+劳动0.5"）解析成 {科目: 节数}。"""
    if header_text is None:
        return {}
    text = str(header_text)
    text = text.replace("\n", "").replace("\r", "").replace("＋", "+")
    text = text.strip()
    if not text:
        return {}
    parts = [p.strip() for p in text.split("+") if p.strip()]
    result = {}
    for p in parts:
        p_clean = re.sub(r"\s+", "", p)
        m = re.match(r"^(.*?)(\d+(?:\.\d+)?)$", p_clean)
        if not m:
            raise ValueError(f"无法解析列头片段: {p!r}（原文: {header_text!r}）")
        name, num_s = m.group(1), m.group(2)
        num = float(num_s)
        if num == int(num):
            num = int(num)
        result[name] = result.get(name, 0) + num
    return result


ROLE_MARKERS = {"数", "语", "体", "音", "美", "科", "班会", "班队"}


def clean_teacher_cell(raw):
    """清洗教师姓名单元格，返回 (base_name, role_suffix, flags dict)。"""
    if raw is None:
        return None, None, {}
    text = str(raw)
    text = text.replace("\n", "").replace("\r", "")
    text = re.sub(r"\s+", "", text)
    text = text.replace("（", "(").replace("）", ")")
    if not text:
        return None, None, {}

    flags = {}
    annotations = re.findall(r"\(([^()]*)\)", text)
    base = re.sub(r"\([^()]*\)", "", text)

    role_suffix = None
    for a in annotations:
        if "班队活动" in a and ("生命" in a or "安全" in a):
            flags["banduihuodong_override"] = True
        elif a.endswith("代"):
            flags["substitute_note"] = a
        elif a in ROLE_MARKERS:
            role_suffix = a
        else:
            flags.setdefault("other_annotations", []).append(a)
    return base, role_suffix, flags


def load_classes():
    """返回 classes: list of dict，每个班一条记录。"""
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)
    sheet_by_stripped_name = {name.strip(): name for name in wb.sheetnames}
    classes = []
    for grade_key in GRADE_ORDER:
        actual_name = sheet_by_stripped_name[grade_key]
        ws = wb[actual_name]
        headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
        # 找到从第3列开始的科目列头解析结果（第1列班级，第2列班主任）
        subject_col_specs = []  # list of (col_idx, {subject: periods})
        for col_idx in range(3, ws.max_column + 1):
            h = headers[col_idx - 1]
            if h is None or not str(h).strip():
                continue
            parsed = parse_header_segment_text(h)
            if parsed:
                subject_col_specs.append((col_idx, parsed))

        for row_idx in range(4, ws.max_row + 1):
            class_no = ws.cell(row=row_idx, column=1).value
            if class_no is None or not str(class_no).strip():
                continue
            try:
                class_no = int(class_no)
            except (TypeError, ValueError):
                continue

            homeroom_raw = ws.cell(row=row_idx, column=2).value
            homeroom_base, homeroom_role, homeroom_flags = clean_teacher_cell(homeroom_raw)

            entries = []
            for col_idx, subj_periods in subject_col_specs:
                cell_raw = ws.cell(row=row_idx, column=col_idx).value
                base, role, flags = clean_teacher_cell(cell_raw)
                if base is None:
                    continue
                primary_subject = next(iter(subj_periods))
                for subj, periods in subj_periods.items():
                    entries.append({
                        "subject": subj,
                        "periods": periods,
                        "teacher_raw": base,
                        "teacher_role_hint": role,
                        "flags": dict(flags),
                        "source_col": col_idx,
                        "col_primary_subject": primary_subject,
                    })

            classes.append({
                "grade_key": grade_key,
                "grade_name": GRADE_NAME[grade_key],
                "campus": CAMPUS[grade_key],
                "class_no": class_no,
                "homeroom_raw": homeroom_base,
                "homeroom_role_hint": homeroom_role,
                "homeroom_flags": homeroom_flags,
                "subjects": entries,
            })
    return classes


# 已确认的同名不同人清单：裸名 -> 用哪个字段区分（primary_subject -> 简称标注）
# 目前只确认了"杨霞"一例（数学老师 vs 体育与健康老师），其余交给 detect_name_collisions 自动检测+人工确认。
CONFIRMED_NAME_COLLISIONS = {
    "杨霞": {"数学": "数", "体育与健康": "体"},
}


def apply_confirmed_disambiguations(classes, confirmed=CONFIRMED_NAME_COLLISIONS):
    """对已确认的同名不同人，把 teacher_raw 改写成"姓名（简称）"，让后续排课把他们当成独立的人。"""
    applied_log = []
    seen = set()
    for c in classes:
        for e in c["subjects"]:
            name = e["teacher_raw"]
            if name in confirmed:
                mapping = confirmed[name]
                marker = mapping.get(e["col_primary_subject"])
                if marker:
                    new_name = f"{name}（{marker}）"
                    log_key = (c["grade_name"], c["class_no"], e["source_col"])
                    if e["teacher_raw"] != new_name and log_key not in seen:
                        seen.add(log_key)
                        applied_log.append(
                            f"{c['grade_name']}{c['class_no']}班：{e['col_primary_subject']} 列的 "
                            f"{e['teacher_raw']} -> {new_name}"
                        )
                    e["teacher_raw"] = new_name
        if c["homeroom_raw"] in confirmed:
            if c["homeroom_role_hint"]:
                new_name = f"{c['homeroom_raw']}（{c['homeroom_role_hint']}）"
                if c["homeroom_raw"] != new_name:
                    applied_log.append(
                        f"{c['grade_name']}{c['class_no']}班：班主任列的 {c['homeroom_raw']} -> {new_name}"
                    )
                c["homeroom_raw"] = new_name
            else:
                applied_log.append(
                    f"【需人工确认】{c['grade_name']}{c['class_no']}班：班主任列是同名冲突姓名"
                    f"「{c['homeroom_raw']}」但没有角色标注，无法自动判断是哪一位，暂保留原名"
                )
    return applied_log


def validate_totals(classes):
    """校验每班总节数是否=26/31，返回问题列表。"""
    problems = []
    for c in classes:
        expected = 26 if c["grade_key"] in ("2026级", "2025级") else 31
        total = sum(e["periods"] for e in c["subjects"])
        if abs(total - expected) > 1e-6:
            problems.append(
                f"{c['grade_name']}{c['class_no']}班（{c['grade_key']}）：期望{expected}节，实际解析出{total}节"
            )
    return problems


def detect_name_collisions(classes):
    """按裸名分组，找出同一裸名在"不同列头"（不同教师-科目组合）下出现明显不兼容学科条线的情况。
    关键：判断依据是"列头的主科目"（每一列头解析出的第一个科目），不是列头里捆绑的所有科目——
    比如"外语3+道德与法治1"是同一个人同一个班捆绑教的，不能因为里面同时含"外语"和"道德与法治"就误判。
    """
    HOMEROOM_PRIMARY = {"语文", "数学"}
    SPECIALIST_PRIMARY = {
        "体育与健康", "音乐", "美术", "科学", "外语", "信息科技",
        "劳动&人工智能", "生命.生态.安全&心理健康",
    }

    name_to_primaries = defaultdict(set)
    name_to_classes = defaultdict(set)
    name_to_role_hints = defaultdict(set)
    for c in classes:
        class_key = (c["grade_key"], c["class_no"])
        for e in c["subjects"]:
            name_to_primaries[e["teacher_raw"]].add(e["col_primary_subject"])
            name_to_classes[e["teacher_raw"]].add(class_key)
            if e["teacher_role_hint"]:
                name_to_role_hints[e["teacher_raw"]].add(e["teacher_role_hint"])

    collisions = []
    for name, primaries in name_to_primaries.items():
        homeroom_hits = primaries & HOMEROOM_PRIMARY
        specialist_hits = primaries & SPECIALIST_PRIMARY
        role_hints = name_to_role_hints.get(name, set())
        suspicious = False
        reason = None
        if homeroom_hits and specialist_hits:
            suspicious = True
            reason = "同一裸名同时出现在语文/数学条线和专职条线的不同列头下"
        elif len(role_hints) >= 2:
            suspicious = True
            reason = "带有多个不同的角色标注（如（数）和（体）同时出现）"
        if suspicious:
            collisions.append({
                "name": name,
                "reason": reason,
                "primary_subjects": sorted(primaries),
                "role_hints": sorted(role_hints),
                "class_count": len(name_to_classes[name]),
            })
    return collisions


if __name__ == "__main__":
    classes = load_classes()
    print(f"共解析出 {len(classes)} 个班")
    problems = validate_totals(classes)
    if problems:
        print("\n=== 节数校验问题 ===")
        for p in problems:
            print(" -", p)
    else:
        print("所有班节数校验通过（=26或31）")

    collisions = detect_name_collisions(classes)
    print("\n=== 疑似同名不同人（自动检测，含已确认的）===")
    for col in collisions:
        print(f" - {col['name']}: {col['reason']}; primary_subjects={col['primary_subjects']}; "
              f"role_hints={col['role_hints']}; 涉及班数={col['class_count']}")

    applied_log = apply_confirmed_disambiguations(classes)
    print(f"\n=== 已应用的同名消歧（共{len(applied_log)}处）===")
    for line in applied_log:
        print(" -", line)

    unconfirmed = [c for c in collisions if c["name"] not in CONFIRMED_NAME_COLLISIONS]
    if unconfirmed:
        print(f"\n=== 仍需人工确认的疑似同名（{len(unconfirmed)}个，未自动消歧）===")
        for col in unconfirmed:
            print(f" - {col['name']}")

    out_path = BASE / "_parsed_classes.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(classes, f, ensure_ascii=False, indent=2)
    print(f"\n已写出 {out_path}")
