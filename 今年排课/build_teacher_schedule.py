# -*- coding: utf-8 -*-
"""把已经排好的班级课表，反过来按教师整理成"每位老师的周课表"，一个教师一个sheet。
sheet顺序按学科+年级分组：语文(一~六年级)、数学(一~六年级)、然后其他学科按《教师与课程安排.xlsx》
里实际出现的科目列顺序，每个学科内部再按年级从小到大排。
"""
import json
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment

from parse_input import load_classes, apply_confirmed_disambiguations, BASE, GRADE_ORDER, GRADE_NAME
from solve import WEEK_LABEL_FIRST, WEEK_LABEL_SECOND

OUTPUT_XLSX = BASE / "教师课程表.xlsx"

DAY_NAMES = ["星期一", "星期二", "星期三", "星期四", "星期五"]
MAX_PERIODS = 7  # 通用网格：周二三到六年级最多7节，其余天/年级用不到的格子留空即可

# 学科大分组顺序：语文、数学放最前（用户明确要求），其余按今年教师安排表里科目列出现的顺序
SUBJECT_GROUP_ORDER = [
    "语文", "数学", "外语", "音乐", "体育与健康", "美术", "科学",
    "劳动&人工智能", "生命.生态.安全&心理健康", "信息科技", "习读本", "道德与法治",
]
GRADE_ORDER_INDEX = {gk: i for i, gk in enumerate(GRADE_ORDER)}  # 2026级=0(一年级)...2021级=5(六年级)


def load_solution():
    sol_path = BASE / "_solution.json"
    with open(sol_path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    solution = {}
    for k, v in raw.items():
        g, n = k.split("|")
        solution[(g, int(n))] = {int(i): slot for i, slot in v.items()}
    return solution


def compute_teacher_buckets(classes):
    """返回 teacher -> (subject分组, grade_key)。
    规则：有语文角色的归语文；否则有数学角色的归数学；否则在他/她所有的(学科,年级)角色里选课时数最多的。
    """
    # teacher -> {(primary_subject, grade_key): 总节数}
    role_periods = defaultdict(lambda: defaultdict(float))
    teacher_min_class = defaultdict(lambda: 9999)
    for c in classes:
        key_grade = c["grade_key"]
        for e in c["subjects"]:
            role_periods[e["teacher_raw"]][(e["col_primary_subject"], key_grade)] += e["periods"]
            teacher_min_class[e["teacher_raw"]] = min(teacher_min_class[e["teacher_raw"]], c["class_no"])

    buckets = {}
    for teacher, roles in role_periods.items():
        chinese_roles = [(s, g) for (s, g) in roles if s == "语文"]
        math_roles = [(s, g) for (s, g) in roles if s == "数学"]
        if chinese_roles:
            buckets[teacher] = chinese_roles[0]
        elif math_roles:
            buckets[teacher] = math_roles[0]
        else:
            best = max(roles.items(), key=lambda kv: kv[1])[0]
            buckets[teacher] = best
    return buckets, teacher_min_class


def sort_key_for_teacher(teacher, buckets, teacher_min_class):
    subject, grade_key = buckets[teacher]
    subject_rank = SUBJECT_GROUP_ORDER.index(subject) if subject in SUBJECT_GROUP_ORDER else len(SUBJECT_GROUP_ORDER)
    grade_rank = GRADE_ORDER_INDEX.get(grade_key, 99)
    return (subject_rank, grade_rank, teacher_min_class[teacher], teacher)


def build_teacher_schedule(classes, all_atoms, solution):
    """返回 teacher -> {(day, period): (class_label, subject)}
    combo（单双周轮换）槎位：两位老师各自只看到自己教的那半，并注明自己是单周还是双周。
    约定与班级课表一致：combo里排在前面的科目=单周，后面的=双周。
    """
    class_label = {(c["grade_key"], c["class_no"]): f"{c['grade_name']}{c['class_no']}班" for c in classes}
    schedule = defaultdict(dict)
    for key, atoms in all_atoms.items():
        if key not in solution:
            continue
        label = class_label[key]
        for idx, slot in solution[key].items():
            atom = atoms[idx]
            day, period = slot // 10, slot % 10
            if atom["kind"] == "combo":
                subj_a, subj_b = atom["subject"].split("/")
                teacher_a, teacher_b = atom["teachers"]
                schedule[teacher_a][(day, period)] = (label, f"{subj_a}（{WEEK_LABEL_FIRST}）")
                schedule[teacher_b][(day, period)] = (label, f"{subj_b}（{WEEK_LABEL_SECOND}）")
            else:
                for t in atom["teachers"]:
                    schedule[t][(day, period)] = (label, atom["subject"])
    return schedule


def write_teacher_workbook(schedule, teacher_order):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    used_names = set()

    for teacher in teacher_order:
        base_name = teacher
        sheet_name = base_name[:31]
        n = 1
        while sheet_name in used_names:
            n += 1
            sheet_name = f"{base_name[:28]}_{n}"
        used_names.add(sheet_name)
        ws = wb.create_sheet(title=sheet_name)

        ws.cell(row=1, column=1, value=f"{teacher} 教师课程表")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        ws.cell(row=2, column=1, value="节次\\星期")
        for day, name in enumerate(DAY_NAMES, start=1):
            ws.cell(row=2, column=day + 1, value=name).alignment = center

        entries = schedule[teacher]
        for period in range(1, MAX_PERIODS + 1):
            row = period + 2
            ws.cell(row=row, column=1, value=f"第{'一二三四五六七'[period-1]}节").alignment = center
            for day in range(1, 6):
                if (day, period) in entries:
                    class_label, subject = entries[(day, period)]
                    ws.cell(row=row, column=day + 1, value=f"{subject}\n{class_label}").alignment = center

        ws.column_dimensions["A"].width = 10
        for col in "BCDEF":
            ws.column_dimensions[col].width = 16
        for period in range(1, MAX_PERIODS + 1):
            ws.row_dimensions[period + 2].height = 30

    wb.save(OUTPUT_XLSX)


if __name__ == "__main__":
    classes = load_classes()
    apply_confirmed_disambiguations(classes)
    from build_atoms import build_all_atoms
    all_atoms, _ = build_all_atoms(classes)
    solution = load_solution()

    schedule = build_teacher_schedule(classes, all_atoms, solution)
    buckets, teacher_min_class = compute_teacher_buckets(classes)
    teacher_order = sorted(schedule.keys(), key=lambda t: sort_key_for_teacher(t, buckets, teacher_min_class))

    write_teacher_workbook(schedule, teacher_order)
    print(f"共{len(schedule)}位教师，已写出 {OUTPUT_XLSX}")

    print("\n分组预览（前30位）：")
    for t in teacher_order[:30]:
        subj, gk = buckets[t]
        print(f"  {t}: {subj} / {GRADE_NAME[gk]}")
